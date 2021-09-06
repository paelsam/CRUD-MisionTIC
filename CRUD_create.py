from openpyxl import load_workbook  # Modulo que nos traera el archivo Excel

def agregar(ruta:int, datos:dict):
    # Importamos el archivo excel a travez de esta variable
    archivo_excel = load_workbook(ruta)
    # Seleccionamos la hoja que queremos sacar
    hoja_datos = archivo_excel['Datos del CRUD']
    # Extraemos la informacion desde la fila A2 hasta la maxima fila, F. 
    # Ahora añadiendo una nueva fila para agregar la nueva informacion
    hoja_datos = hoja_datos['A2':'F' + str(hoja_datos.max_row + 1)]
    hoja = archivo_excel.active # Variable que activira la edicion del archivo Excel

    titulo, descripcion, estado, fecha_inicio, fecha_finalizado = 2,3,4,5,6
    
    for i in hoja_datos:
        if not(isinstance(i[0].value, int)):
            identificador = i[0].row
            hoja.cell(row=identificador, column=1).value = identificador + 1
            hoja.cell(row=identificador, column=titulo).value=datos['Titulo']
            hoja.cell(row=identificador, column=descripcion).value=datos['Descripcion']
            hoja.cell(row=identificador, column=estado).value=datos['Estado']
            hoja.cell(row=identificador, column=fecha_inicio).value=datos['Fecha Inicio']
            hoja.cell(row=identificador, column=fecha_finalizado).value=datos['Fecha Finalizacion']
            break
    archivo_excel.save(ruta)
    return "Succeful Operation"


            
