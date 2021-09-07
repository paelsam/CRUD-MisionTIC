from openpyxl import load_workbook  # Modulo que nos traera el archivo Excel

def borrar(ruta, identificador): # Parametros: Ruta del archivo y El ID a borrar
    # Importamos el archivo excel a travez de esta variable
    archivo_excel = load_workbook(ruta)
    # Seleccionamos la hoja que queremos sacar
    hoja_datos = archivo_excel['Datos del CRUD']
    # Extraemos la informacion desde la fila A2 hasta la maxima fila, F
    hoja_datos = hoja_datos['A2':'F' + str(hoja_datos.max_row)]
    # La funcion 'active' nos ayuda a editar el documento Excel
    hoja = archivo_excel.active # Variable que activira la edicion del archivo Excel

    titulo, descripcion, estado, fecha_inicio, fecha_finalizado, encontro = 2,3,4,5,6,False
    
    for i in hoja_datos:
        if i[0].value == identificador:
            fila = i[0].row # Fila a eliminar
            encontro = True
            hoja.delete_rows(fila)
    archivo_excel.save(ruta)
    if encontro == False:
        print("ERROR: No existe un archivo con ese ID")
    else:
        return "Ha sido borrado con Exito!"

    
    
    