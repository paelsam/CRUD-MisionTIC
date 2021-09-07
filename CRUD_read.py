from datetime import datetime  # Aun no los necesitamos, asi que ignorenlo XD
from openpyxl import load_workbook  # Modulo que nos traera el archivo Excel


def leer(ruta: str, extraer: str):
    # Importamos el archivo excel a travez de esta variable
    archivo_excel = load_workbook(ruta)
    # Seleccionamos la hoja que queremos sacar
    hoja_datos = archivo_excel['Datos del CRUD']
    # Extraemos la informacion desde la fila A2 hasta la maxima fila, F
    hoja_datos = hoja_datos['A2':'F' + str(hoja_datos.max_row)]

    info = {}  # Diccionario donde se guardaran los datos leidos desde el Excel

    for i in hoja_datos:  # Recorremos la variable hoja_datos que es una tupla y sus llaves son las filas del documento Excel
        if isinstance(i[0].value, int):
            info.setdefault(i[0].value, {  # Guardamos los valores las filas en el diccionario
                'tarea': i[1].value,
                'descripcion': i[2].value,
                'estado': i[3].value,
                'fecha de inicio': i[4].value,
                'fecha de finalizacion': i[5].value
            })

    def filtrar_estado(info, filtro):  # Para filtrar un archivo por estados.
        for i in info:  # For que tomara los datos del Excel
            if info[i]['estado'] == filtro:  # Condicional que filtrara el estado correspondiente
                print('*********Tarea**********')
                result = print(
                        'ID:', str(i),
                        '\nTitulo:', str(info[i]['tarea']),
                        '\nDescripcion:', str(info[i]['descripcion']),
                        '\nEstado:', str(info[i]['estado']),
                        '\nFecha de inicio:', str(info[i]['fecha de inicio']),
                        '\nFecha de finalizacion', str(info[i]['fecha de finalizacion'])
                        )
                print('************************\n')
        return result

    # Esto es por si queremos filtrar, mostrara todo si escribimos el estado que queramos filtrar.
    if not(extraer == 'todo'):
        print(filtrar_estado(info, extraer))
    else:
        for i in info:  # Si no, mostrara todo con esto :3
            print('*********Tarea**********')
            print(
                'ID:', str(i),
                '\nTitulo:', str(info[i]['tarea']),
                '\nDescripcion:', str(info[i]['descripcion']),
                '\nEstado:', str(info[i]['estado']),
                '\nFecha de inicio:', str(info[i]['fecha de inicio']),
                '\nFecha de finalizacion', str(info[i]['fecha de finalizacion']))
            print('************************\n')
    return "Succeful Process"



