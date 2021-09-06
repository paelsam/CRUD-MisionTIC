from openpyxl import load_workbook  # Modulo que nos traera el archivo Excel


def actualizar(ruta: str, identificador: int, datos_actualizados: dict):
    # Importamos el archivo excel a travez de esta variable
    archivo_excel = load_workbook(ruta)
    # Seleccionamos la hoja que queremos sacar
    hoja_datos = archivo_excel['Datos del CRUD']
    # Extraemos la informacion desde la fila A2 hasta la maxima fila, F
    hoja_datos = hoja_datos['A2':'F' + str(hoja_datos.max_row)]
    # La funcion 'active' nos ayuda a editar el documento Excel
    hoja = archivo_excel.active

    # Indicaremos las columnas de nuestro documento
    titulo = 2
    descripcion = 3
    estado = 4
    fecha_inicio = 5
    fecha_finalizado = 6
    encontro = False

    for i in hoja_datos:
        if i[0].value == identificador:
            # La variable fila capturara la fila de su ID
            fila = i[0].row
            # La Flag Variable para mostrar que si encontro el ID
            encontro = True
            for d in datos_actualizados:
                # Condicional para actualizar los valores
                # Si el espacio esta vacio, no lo actualiza
                if d == 'Titulo' and not(datos_actualizados[d] == ''):
                    # La funcion 'cell' nos deja acceder a la celda.
                    # Teniendo en cuenta la direccion de su fila y columna.
                    hoja.cell(
                        row=fila, column=titulo).value = datos_actualizados[d]
                elif d == 'Descripcion' and not(datos_actualizados[d] == ''):
                    hoja.cell(
                        row=fila, column=descripcion).value = datos_actualizados[d]
                elif d == 'Estado' and not(datos_actualizados[d] == ''):
                    hoja.cell(
                        row=fila, column=estado).value = datos_actualizados[d]
                elif d == 'Fecha Inicio' and not(datos_actualizados[d] == ''):
                    hoja.cell(
                        row=fila, column=fecha_inicio).value = datos_actualizados[d]
                elif d == 'Fecha Finalizado' and not(datos_actualizados[d] == ''):
                    hoja.cell(
                        row=fila, column=fecha_finalizado).value = datos_actualizados[d]
    archivo_excel.save(ruta)  # Guardamos los cambios
    if encontro == False:  # Por si el ID no existe
        print("No se encontro el ID especificado. Intentalo de Nuevo")
    return

