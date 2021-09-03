from datetime import datetime # Aun no los necesitamos, asi que ignorenlo XD
from openpyxl import load_workbook # Modulo que nos traera el archivo Excel

def leer(ruta:str, extraer:str):
    archivo_excel = load_workbook(ruta) # Importamos el archivo excel a travez de esta variable
    hoja_datos = archivo_excel['Datos del CRUD'] # Seleccionamos la hoja que queremos sacar
    hoja_datos=hoja_datos['A2':'F' + str(hoja_datos.max_row)] # Extraemos la informacion desde la fila A2 hasta la maxima fila, F

    info = {} # Diccionario donde se guardaran los datos leidos desde el Excel

    for i in hoja_datos: # Recorremos la variable hoja_datos que es una tupla y sus llaves son las filas del documento Excel
        if isinstance(i[0].value, int):
            info.setdefault(i[0].value, { # Guardamos los valores las filas en el diccionario
            'tarea': i[1].value,
            'descripcion': i[2].value,
            'estado': i[3].value,
            'fecha de inicio': i[4].value,
            'fecha de finalizacion': i[5].value
            })

    def filtrar_estado(info, filtro): # Para filtrar un archivo por estados.
        for i in info: # For que tomara los datos del Excel
            if info[i]['estado']==filtro: # Condicional que filtrara el estado correspondiente
                aux =print('ID: ', str(i), '\nTitulo:', str(info[i]['tarea']),
                '\nDescripcion:', str(info[i]['descripcion']), '\nEstado:',
                str(info[i]['estado']), '\nFecha de inicio:', str(info[i]['fecha de inicio']),
                '\nFecha de finalizacion', str(info[i]['fecha de finalizacion']))
        return aux


    if not(extraer=='todo'): # # Esto es por si queremos filtrar, mostrara todo si escribimos el estado que queramos filtrar.
        print(filtrar_estado(info, extraer))
    else: #
        for i in info: # Si no, mostrara todo con esto :3
            print('*********Tarea**********')
            print('ID: ', str(i), '\nTitulo:', str(info[i]['tarea']),
            '\nDescripcion:', str(info[i]['descripcion']), '\nEstado:',
            str(info[i]['estado']), '\nFecha de inicio:', str(info[i]['fecha de inicio']),
            '\nFecha de finalizacion', str(info[i]['fecha de finalizacion']))
            print('************************')
    return



rut = r"C:\Users\kinmi\OneDrive\Escritorio\CRUD\docs\Base CRUD.xlsx" # Si lo van a probar, tienen que cambiar la direccion de su archivo de Excel
# Recuerda que esta debe tener todas las cosas especificadas en el ejemplo de la clase 1.
leer(rut, "todo")
